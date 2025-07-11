import openpyxl as xl
from openpyxl.chart import BarChart, Reference

def process_wb(filename):
    wb = xl.load_workbook(filename) #access file
    sheet = wb['Sheet1'] #access sheet in file  
    #cell = sheet['a1'] #access cell in sheet 


    for row in range (2, sheet.max_row + 1):
        cell = sheet.cell(row, 3)
        corrected_price = cell.value * 0.9
        cell1 = sheet.cell(row, 4)
        cell1.value = corrected_price

    values = Reference(sheet, 
        min_row=2, 
        max_row=sheet.max_row,  
        min_col=4, 
        max_col=4)

    chart = BarChart()
    chart.add_data(values)
    sheet.add_chart(chart, 'e2')
    wb.save(filename)

process_wb('transactions.xlsx')